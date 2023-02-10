

import re
import requests
from openpyxl import load_workbook
from openpyxl import Workbook

'''

test_string = """
aaa@bbb.com
123@abc.co.kr
test@hello.kr
ok@ok.co.kr
ok@ok.co.kr
no.co.kr
no.kr
"""

result = re.findall(r'[\w\.-]+@[\w\.-]+',test_string)       # 이메일 형식으로 된 것 확인

results = list(set(results))                                # 중복 제거 코드

print(results)

'''

url = 'https://news.v.daum.net/v/20211129144552297'

headers = {
    'User-Agent':'Mozilla/5.0',
    'Content-Type':'Text/html;charset=utf-8'
}

response = requests.get(url,headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+',response.text)

results = list(set(results))

print(results)

try:
    wb = load_workbook(r'/Users/chung_sungwoong/Desktop/Practice/Practice_Python/email.xlsx',data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save(r'/Users/chung_sungwoong/Desktop/Practice/Practice_Python/email.xlsx')