"""
엑셀 파일 불러오기
"""

from openpyxl import load_workbook

load_wb = load_workbook(r"/Users/chung_sungwoong/Desktop/Practice/Practice_Python/LIst_of_Certificate.xlsx")
load_ws = load_wb.active

name_list = []
birthday_list = []
ho_list = []

for i in range(1,load_ws.max_row+1):
    name_list.append(load_ws.cell(i,1).value)
    birthday_list.append(load_ws.cell(i,2).value)
    ho_list.append(load_ws.cell(i,3).value)

print(name_list)
print(birthday_list)
print(ho_list)